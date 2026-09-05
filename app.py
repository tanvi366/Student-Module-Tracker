from flask import Flask, render_template, request, jsonify, make_response, send_file
import sqlite3
import csv
import io
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
app = Flask(__name__)

DATABASE = "career_tracker.db"

def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()

    conn.execute("""
        CREATE TABLE IF NOT EXISTS applications(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT NOT NULL,
            role TEXT NOT NULL,
            location TEXT,
            status TEXT NOT NULL DEFAULT 'Applied',
            application_date TEXT NOT NULL,
            job_url TEXT,
            notes TEXT
        )
    """)
    try: 
         conn.execute(" ALTER TABLE applications ADD COLUMN deadline TEXT")
    except sqlite3.OperationalError:
        pass
    conn.commit()
    conn.close()


@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/applications", methods=["GET"])
def get_applications():
    conn = get_db_connection()
    applications = conn.execute("""
            SELECT * FROM applications ORDER BY application_date DESC
    """).fetchall()

    conn.close()

    return jsonify([
        dict(application)
        for application in applications
    ])

@app.route("/api/applications", methods=["POST"])
def create_application():
    data = request.get_json()

    company = data.get("company")
    if not company or not company.strip():
        return "Company name is required", 400

    role = data.get("role")
    if not role or not role.strip():
        return "Role is required", 400
    location = data.get("location")
    status = data.get("status", "Applied")
    application_date = data.get("application_date")
    deadline = data.get("deadline")
    job_url = data.get("job_url")
    notes = data.get("notes")

    conn = get_db_connection()

    cursor = conn.execute("""
        INSERT INTO applications(company,role,location,status,application_date,job_url,notes,deadline) VALUES(?,?,?,?,?,?,?,?)""",
        (company,role,location,status,application_date,job_url,notes,deadline))

    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return jsonify({
        "message": "Application created",
        "id": new_id
    }), 201

@app.route("/api/applications/<int:id>", methods=["DELETE"])
def delete_application(id):
    conn = get_db_connection()
    conn.execute("DELETE FROM applications WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return jsonify({"message": "Application deleted"})

@app.route("/api/applications/<int:id>", methods=["PUT"])
def update_applications(id):
    data = request.get_json()
    conn = get_db_connection()
    conn.execute("""
        UPDATE applications SET company=?, role=?, location=?, status=?, application_date=?, job_url=?, notes=?, deadline=? WHERE id=?
    """,
    (data["company"], data["role"], data["location"], data["status"], data["application_date"], data["job_url"], data["notes"], data["deadline"], id
    ))
    conn.commit()
    conn.close()
    return jsonify({"message": "Application updated"})

@app.route("/export/csv")
def export_csv():
    conn = get_db_connection()
    applications = conn.execute("""SELECT * FROM applications ORDER BY application_date DESC""").fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Company", "Role", "Location", "Status", "Application Date", "Deadline", "Job URL", "Notes" ])
    for application in applications:
        writer.writerow([application["company"], application["role"], application["location"], application["status"], application["application_date"], application["deadline"], application["job_url"], application["notes"]])
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = ("attachment; filename=career_tracker_export.csv")
    response.headers["Content-Type"] = "text/csv"
    return response

@app.route("/export/excel")
def export_excel():
    conn = get_db_connection()
    applications = conn.execute("""SELECT * FROM applications ORDER BY application_date DESC""").fetchall()
    conn.close()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Applications"
    sheet["A1"] = "Career Tracker Report"

    sheet["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A1:H1")
    headers =["Company", "Role", "Location", "Status", "Application Date", "Deadline", "Job URL", "Notes" ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column = column)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")
        cell.alignment = Alignment(horizontal = "center")
    for row, application in enumerate(applications, start=4):
        sheet.cell(row=row, column=1, value=application["company"])
        sheet.cell(row=row, column=2, value=application["role"])
        sheet.cell(row=row, column=3, value=application["location"])
        sheet.cell(row=row, column=4, value=application["status"])
        sheet.cell(row=row, column=5, value=application["application_date"])
        sheet.cell(row=row, column=6, value=application["deadline"])
        sheet.cell(row=row, column=7, value=application["job_url"])
        sheet.cell(row=row, column=8, value=application["notes"])

    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 22
    sheet.column_dimensions["F"].width = 22
    sheet.column_dimensions["G"].width = 35
    sheet.column_dimensions["H"].width = 40

    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Career Tracker Summary"
    summary["A1"].font = Font(size=20, bold=True)
    summary["A3"] = "Status"
    summary["B3"] = "Applications"
    summary["A3"].font = Font(bold=True)
    summary["B3"].font = Font(bold=True)
    statuses = ["Applied", "Online Assessment", "Interview", "Offer", "Rejected"]
    for row, status in enumerate(statuses, start=4):
        summary.cell(row=row, column=1, value=status)
        count=sum(1 for application in applications if application["status"]==status)
        summary.cell(row=row, column=2, value=count)
    summary["A9"] = "Total Applications"
    summary["B9"] = len(applications)

    summary["A9"].font = Font(bold=True)
    summary["B9"].font = Font(bold=True)

    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 20
    
    sheet.freeze_panes = "A4"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="career_tracker_report.xlsx", mimetype=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))

if __name__ == "__main__":
    init_db()
    #alter_db()
    app.run()
